"""
Python code for lagging dataset so it may be used for LSTM training
"""
#IMPORTS
import numpy as np
import pandas as pd
import cv2
import os
from openpyxl import load_workbook

def lag_dataset(data, n_in=1, n_out=1, dropnan=True):
  """
  @param data: Data being lagged
  @param n_in: Number of lags for input
  @param n_out: number of lags for output 
  @param dropnan: if true, drops NAN values
  @return df_lagged: dataset after lagging has been applied
  """
  #extract number of variables 
  n_vars = 1 if type(data) is list else data.shape[1]
  df = pd.DataFrame(data)
  cols, names = list(), list()
  # input sequence 
  for i in range(n_in, 0, -1):
    #shift the column the specified amount of times 
    cols.append(df.shift(i))
    names += [('var%d(t-%d)' % (j+1, i)) for j in range(n_vars)]
  # output sequence 
  for i in range(0, n_out):
    cols.append(df.shift(-i))
    if i == 0:
      names += [('var%d(t)' % (j+1)) for j in range(n_vars)]
    else:
      names += [('var%d(t+%d)' % (j+1, i)) for j in range(n_vars)]
  # put it all together
  df_lagged = pd.concat(cols, axis=1)
  df_lagged .columns = names
  # drop rows with NaN values
  if dropnan:
    df_lagged .dropna(inplace=True)
  return df_lagged 

def preprocess_df(df):
  """"
  @param: original dataframe 
  @return df: preprocessed dataset
  """
  df.drop(['frame'],axis=1,inplace=True)
  return df

#import excel file to store final data
book = load_workbook('LSTM_DATA2.xlsx')
writer=pd.ExcelWriter('LSTM_DATA2.xlsx',engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

for filename in os.listdir('.'):
    #loop through all files in dataset 
    df=pd.read_excel(filename,engine='openpyxl')
    #preprocess the dataset
    df_preprocessed=preprocess_df(df)
    #extract values from dataset 
    values = df_preprocessed.values
    values=values.astype('float32')
    #lag the dataset (creating window)
    df_lagged = lag_dataset(values, 200, 1)
    #write data into excel file 
    df_lagged.to_excel(writer, startrow = writer.sheets['Sheet1'].max_row,index = False, header = False)
    writer.save()


